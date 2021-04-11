import pandas
import openpyxl
vykazy = pandas.read_csv("vykazy.csv")
info_pro_zakaznika = vykazy.groupby("project")["hours"].sum()
# ulož tabulku s celkovými počty hodin z příkladu 27 jako excel soubor
# info_pro_zakaznika.to_excel("info_pro_zakaznika.xlsx")
# jde tam taky dát přímo cesta, místo toho názvu v uvozovkách:C:\Users\ksodomkova\PycharmProjects\python-012021\6\ukoly6\info_pro_zakaznika.xlsx

# zkus data z Excelu naopak načíst pomocí funkce read_excel() a ověř, že se soubor načetl v pořádku
zkouska = pandas.read_excel("info_pro_zakaznika.xlsx")

# rozšířené zadání
from openpyxl import Workbook
from openpyxl.styles import PatternFill, GradientFill

wb = Workbook()
ws1 = wb.active
ws1.title = "rozvrh"

rozvrh_hodin = [
  ["Anglický jazyk", "Přírodopis", "Dějepis", "Matematika", "Oběd", "Tělesná výchova", "Tělesná výchova", ],
  ["Občanská výchova", "Hudební výchova", "Matematika", "Oběd", "Výtvarná výchova", "Dějepis", ],
  ["Matematika", "Chemie", "Přírodopis", "Fyzika", "Oběd", "Zeměpis", ],
  ["Fyzika", "Anglický jazyk", "Matematika", "Český jazyk", "Dějepis", "Oběd", ],
  ["Český jazyk", "Zeměpis", "Český jazyk", "Výtvarná výchova", "Oběd", ]
]
radek = 1
for den in rozvrh_hodin:
  sloupec = 1
  for predmet in den:
    bunka = ws1.cell(radek, sloupec, predmet)
    if bunka.value == "Anglický jazyk" or bunka.value == "Český jazyk":
      prechodova_barva = GradientFill(stop=("000000", "FFFFFF"))
      bunka.fill = prechodova_barva
    if bunka.value == "Matematika":
      ruzova_barva = PatternFill("solid", fgColor="00FF99CC")
      bunka.fill = ruzova_barva
    if radek == 1:
      modra_barva = PatternFill("solid", fgColor="00CCCCFF")
      bunka.fill = modra_barva
    sloupec += 1
  radek += 1

wb.save(filename="rozvrh_hodin.xlsx")